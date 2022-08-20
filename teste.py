from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from time import sleep
import openpyxl

# localizando o arquivo de dados
caminho = "C:\\Users\\pigsr\\PycharmProjects\\Desafio MercaFacil\\CT 3\\massa_de_dados\\dados.xlsx"

# lendo o arquivo de daodos
massa_dados = openpyxl.load_workbook(caminho)
sheet_obj = massa_dados.active

link = sheet_obj.cell(row = 2, column = 1)
cnpj = sheet_obj.cell(row = 2, column = 2)
razao_social = sheet_obj.cell(row = 2, column = 3)
celular = sheet_obj.cell(row = 2, column = 4)
tel_contato = sheet_obj.cell(row = 2, column = 5)
email = sheet_obj.cell(row = 2, column = 6)
cep = sheet_obj.cell(row = 2, column = 7)
numero = sheet_obj.cell(row = 2, column = 8)

# Acessando a pagina de cadastro CNPJ

nav = webdriver.Chrome()
nav.get(link.value)
nav.maximize_window()
sleep(3)
nav.find_element(By.XPATH,'/html/body/div/div[2]/div/main/div/section/div/div[2]/button').click()
sleep(3)
nav.find_element(By.ID,'input-19').send_keys(cnpj.value)
sleep(3)
nav.find_element(By.XPATH,'/html/body/div/div[2]/div/main/div/div/div[2]/div/div/div[2]/button').click()
sleep(3)

# validando os campos de cadastro

nav.find_element(By.ID,'input-46').send_keys(Keys.TAB)
sleep(3)
nav.find_element(By.ID,'input-46').send_keys(razao_social.value)
sleep(2)
nav.find_element(By.ID,'input-50').send_keys(celular.value)
sleep(3)
nav.find_element(By.ID,'input-50').send_keys(Keys.TAB)
sleep(3)
nav.find_element(By.ID,'input-54').send_keys(Keys.TAB)
sleep(2)
nav.find_element(By.ID,'input-54').send_keys(tel_contato.value)
sleep(2)
nav.find_element(By.ID,'input-58').send_keys(email.value)
sleep(2)
nav.find_element(By.ID,'input-75').send_keys(Keys.TAB)
sleep(2)
nav.find_element(By.ID,'input-75').send_keys(cep.value)
sleep(2)
nav.find_element(By.ID,'input-102').send_keys(Keys.DOWN)
sleep(2)
nav.find_element(By.ID,'input-102').send_keys(numero.value)
sleep(2)
nav.find_element(By.ID,'input-102').send_keys(Keys.TAB)
sleep(2)
nav.find_element(By.ID,'input-102').send_keys(Keys.PAGE_DOWN)
sleep(3)

# marcando os termos de adesão

nav.find_element(By.XPATH,'/html/body/div/div[2]/div/main/div/div/div/div[2]/div/div/div/div[9]/div[1]').click()
sleep(3)
nav.find_element(By.XPATH,'/html/body/div/div[2]/div[3]/div/div/div[2]/button[1]').click()
sleep(3)
nav.delete_all_cookies()
sleep(2)
nav.close()